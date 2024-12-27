"""Abstract interface for document loader implementations."""

import os
from typing import Optional

import pandas as pd
from openpyxl import load_workbook

from dify_rag.extractor.extractor_base import BaseExtractor
from dify_rag.extractor.html_extractor import HtmlExtractor
from dify_rag.extractor.utils import get_encoding
from dify_rag.models.document import Document


class ExcelExtractor(BaseExtractor):
    """Load Excel files.


    Args:
        file_path: Path to the file to load.
    """

    def __init__(
        self,
        file_path: str,
        file_name: Optional[str] = None,
        encoding: Optional[str] = None,
        autodetect_encoding: bool = False,
    ):
        """Initialize with file path."""
        self._file_path = file_path
        self._encoding = encoding or get_encoding(file_path)
        self._autodetect_encoding = autodetect_encoding
        self._file_name = file_name
        if file_name:
            self._file_name = os.path.basename(file_name).split(".")[0]

    def extract(self) -> list[Document]:
        """Load from Excel file in xls or xlsx format using Pandas and openpyxl."""
        documents = []
        file_extension = os.path.splitext(self._file_path)[-1].lower()

        if file_extension == ".xlsx":
            wb = load_workbook(self._file_path, data_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                data = sheet.values
                try:
                    cols = next(data)
                except StopIteration:
                    continue
                df = pd.DataFrame(data, columns=cols)
                df.dropna(how="all", inplace=True)

        elif file_extension == ".xls":
            excel_file = pd.ExcelFile(self._file_path, engine="xlrd")
            for sheet_name in excel_file.sheet_names:
                df = excel_file.parse(sheet_name=sheet_name)
                df.dropna(how="all", inplace=True)

        else:
            raise ValueError(f"Unsupported file extension: {file_extension}")

        html_content = df.to_html(index=False)
        extractor = HtmlExtractor(file=html_content)

        docs = extractor.extract()
        documents.extend(docs)

        return documents
